import telebot
from telebot import types
import database as db
import config
import logging
import io

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

bot = telebot.TeleBot(config.BOT_TOKEN)

user_states = {}
STEPS = ["name", "phone", "age", "course"]

def is_admin(user_id):
    return user_id in config.ADMIN_IDS

def main_menu(user_id):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    if is_admin(user_id):
        markup.add("👥 O'quvchilar ro'yxati", "✅ Kutayotganlar")
        markup.add("📢 Xabar yuborish", "📚 Kurslar")
        markup.add("📅 Jadval boshqaruv", "📤 Excel eksport")
    else:
        markup.add("📝 Ro'yxatdan o'tish", "📅 Dars jadvali")
        markup.add("🏆 Kurs narxlari", "ℹ️ Ma'lumot")
    return markup

def course_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    courses = db.get_all_courses()
    for course in courses:
        markup.add(f"📚 {course['name']}")
    markup.add("🔙 Bekor qilish")
    return markup

@bot.message_handler(commands=["start"])
def start(message):
    user_id = message.from_user.id
    db.add_user_if_not_exists(user_id, message.from_user.username)
    name = message.from_user.first_name or "Foydalanuvchi"
    bot.send_message(
        user_id,
        f"👋 Salom, *{name}*!\n\n🏫 *Opus o'quv markaziga xush kelibsiz!*\n\nQuyidagi menyularni tanlang 👇",
        parse_mode="Markdown",
        reply_markup=main_menu(user_id)
    )

@bot.message_handler(func=lambda m: m.text == "📝 Ro'yxatdan o'tish")
def start_registration(message):
    user_id = message.from_user.id
    if db.is_already_registered(user_id):
        status = db.get_student_status(user_id)
        status_text = {
            "pending": "⏳ Arizangiz ko'rib chiqilmoqda...",
            "approved": "✅ Siz allaqachon tasdiqlangansiz!",
            "rejected": "❌ Arizangiz rad etilgan."
        }.get(status, "Noma'lum holat")
        bot.send_message(user_id, status_text)
        return
    user_states[user_id] = {"step": "name", "data": {}}
    bot.send_message(
        user_id,
        "📝 *Ro'yxatdan o'tish 1/4*\n\nTo'liq ism-familiyangizni kiriting:",
        parse_mode="Markdown",
        reply_markup=types.ReplyKeyboardMarkup(resize_keyboard=True).add("🔙 Bekor qilish")
    )

@bot.message_handler(func=lambda m: m.from_user.id in user_states and user_states[m.from_user.id].get("step") in STEPS)
def handle_registration_steps(message):
    user_id = message.from_user.id
    text = message.text.strip()
    if text == "🔙 Bekor qilish":
        user_states.pop(user_id, None)
        bot.send_message(user_id, "❌ Bekor qilindi.", reply_markup=main_menu(user_id))
        return
    state = user_states[user_id]
    step = state["step"]
    if step == "name":
        if len(text) < 3:
            bot.send_message(user_id, "⚠️ Juda qisqa. To'liq ism-familiya kiriting:")
            return
        state["data"]["name"] = text
        state["step"] = "phone"
        phone_btn = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        phone_btn.add(types.KeyboardButton("📱 Telefon raqamimni yuborish", request_contact=True))
        phone_btn.add("🔙 Bekor qilish")
        bot.send_message(user_id, "📝 *2/4*\n\nTelefon raqamingizni yuboring:", parse_mode="Markdown", reply_markup=phone_btn)
    elif step == "phone":
        if not (text.startswith("+") or text.startswith("9")):
            bot.send_message(user_id, "⚠️ Noto'g'ri format. Masalan: +998901234567")
            return
        state["data"]["phone"] = text
        state["step"] = "age"
        bot.send_message(user_id, "📝 *3/4*\n\nYoshingiz yoki sinfingizni kiriting:", parse_mode="Markdown",
                         reply_markup=types.ReplyKeyboardMarkup(resize_keyboard=True).add("🔙 Bekor qilish"))
    elif step == "age":
        state["data"]["age"] = text
        state["step"] = "course"
        courses = db.get_all_courses()
        if not courses:
            bot.send_message(user_id, "⚠️ Hozircha kurslar mavjud emas.")
            user_states.pop(user_id, None)
            return
        bot.send_message(user_id, "📝 *4/4*\n\nQaysi kursga qiziqasiz?", parse_mode="Markdown", reply_markup=course_keyboard())
    elif step == "course":
        course_name = text.replace("📚 ", "")
        courses = [c["name"] for c in db.get_all_courses()]
        if course_name not in courses:
            bot.send_message(user_id, "⚠️ Ro'yxatdan kurs tanlang:")
            return
        state["data"]["course"] = course_name
        data = state["data"]
        db.save_student(user_id=user_id, name=data["name"], phone=data["phone"], age=data["age"], course=data["course"])
        user_states.pop(user_id, None)
        bot.send_message(user_id,
            f"✅ *Arizangiz qabul qilindi!*\n\n👤 {data['name']}\n📱 {data['phone']}\n🎂 {data['age']}\n📚 {data['course']}\n\n⏳ Admin ko'rib chiqishi kutilmoqda...",
            parse_mode="Markdown", reply_markup=main_menu(user_id))
        for admin_id in config.ADMIN_IDS:
            try:
                markup = types.InlineKeyboardMarkup()
                markup.add(
                    types.InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_{user_id}"),
                    types.InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_{user_id}")
                )
                bot.send_message(admin_id,
                    f"🔔 *Yangi ariza!*\n\n👤 {data['name']}\n📱 {data['phone']}\n🎂 {data['age']}\n📚 {data['course']}\n🆔 `{user_id}`",
                    parse_mode="Markdown", reply_markup=markup)
            except Exception as e:
                logger.error(f"Admin {admin_id}: {e}")

@bot.message_handler(content_types=["contact"])
def handle_contact(message):
    user_id = message.from_user.id
    if user_id in user_states and user_states[user_id].get("step") == "phone":
        phone = message.contact.phone_number
        if not phone.startswith("+"): phone = "+" + phone
        user_states[user_id]["data"]["phone"] = phone
        user_states[user_id]["step"] = "age"
        bot.send_message(user_id, f"✅ {phone}\n\n📝 *3/4*\n\nYoshingiz yoki sinfingizni kiriting:", parse_mode="Markdown",
                         reply_markup=types.ReplyKeyboardMarkup(resize_keyboard=True).add("🔙 Bekor qilish"))

@bot.message_handler(func=lambda m: m.text == "📅 Dars jadvali" and not is_admin(m.from_user.id))
def show_schedule(message):
    schedules = db.get_all_schedules()
    if not schedules:
        bot.send_message(message.chat.id, "📭 Hozircha dars jadvali mavjud emas.")
        return
    from collections import defaultdict
    by_course = defaultdict(list)
    for s in schedules:
        by_course[s['course']].append(s)
    text = "📅 *Dars jadvali:*\n\n"
    for course, items in by_course.items():
        text += f"━━━━━━━━━━━━━━\n📚 *{course}*\n"
        for item in items:
            text += f"👨‍🏫 {item['teacher']} | 🎯 {item['level']}\n📆 {item['days']} | 🕐 {item['time']}\n\n"
    bot.send_message(message.chat.id, text, parse_mode="Markdown")

@bot.message_handler(func=lambda m: m.text == "🏆 Kurs narxlari")
def show_prices(message):
    courses = db.get_courses_with_price()
    if not courses:
        bot.send_message(message.chat.id, "📭 Narxlar hali kiritilmagan.")
        return
    text = "🏆 *Kurs narxlari:*\n\n"
    for c in courses:
        price = f"{c.get('price', 0) or 0:,} so'm" if c.get('price') else "Aniqlanmagan"
        text += f"📚 *{c['name']}* — {price}\n"
    bot.send_message(message.chat.id, text, parse_mode="Markdown")

@bot.message_handler(func=lambda m: m.text == "ℹ️ Ma'lumot")
def info(message):
    bot.send_message(message.from_user.id,
        "🏫 *Opus o'quv markazi*\n\nBizning markazda sifatli ta'lim beriladi.\nRo'yxatdan o'tish uchun '📝 Ro'yxatdan o'tish' tugmasini bosing.\n\n📞 Aloqa: @admin_username",,
        parse_mode="Markdown")

@bot.message_handler(func=lambda m: m.text == "👥 O'quvchilar ro'yxati" and is_admin(m.from_user.id))
def students_list(message):
    students = db.get_approved_students()
    if not students:
        bot.send_message(message.chat.id, "📭 Tasdiqlangan o'quvchilar yo'q.")
        return
    text = f"👥 *Tasdiqlangan ({len(students)} ta):*\n\n"
    for i, s in enumerate(students, 1):
        text += f"{i}. {s['name']} | {s['course']} | {s['phone']}\n"
    bot.send_message(message.chat.id, text, parse_mode="Markdown")

@bot.message_handler(func=lambda m: m.text == "✅ Kutayotganlar" and is_admin(m.from_user.id))
def pending_list(message):
    students = db.get_pending_students()
    if not students:
        bot.send_message(message.chat.id, "📭 Kutayotgan arizalar yo'q.")
        return
    for s in students:
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_{s['user_id']}"),
            types.InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_{s['user_id']}")
        )
        bot.send_message(message.chat.id,
            f"👤 *{s['name']}*\n📱 {s['phone']}\n🎂 {s['age']}\n📚 {s['course']}\n🆔 `{s['user_id']}`",
            parse_mode="Markdown", reply_markup=markup)

@bot.callback_query_handler(func=lambda c: c.data.startswith("approve_") or c.data.startswith("reject_"))
def handle_decision(call):
    if not is_admin(call.from_user.id):
        bot.answer_callback_query(call.id, "❌ Ruxsat yo'q!")
        return
    action, student_id = call.data.split("_", 1)
    student_id = int(student_id)
    student = db.get_student(student_id)
    if not student:
        bot.answer_callback_query(call.id, "❌ Topilmadi.")
        return
    if action == "approve":
        db.update_student_status(student_id, "approved")
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
        bot.answer_callback_query(call.id, "✅ Tasdiqlandi!")
        bot.send_message(call.message.chat.id, f"✅ {student['name']} tasdiqlandi.")
        try:
            bot.send_message(student_id, "🎉 *Tabriklaymiz!*\n\nArizangiz tasdiqlandi! Tez orada bog'lanamiz. 📞", parse_mode="Markdown")
        except: pass
    elif action == "reject":
        db.update_student_status(student_id, "rejected")
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
        bot.answer_callback_query(call.id, "❌ Rad etildi.")
        bot.send_message(call.message.chat.id, f"❌ {student['name']} rad etildi.")
        try:
            bot.send_message(student_id, "😔 *Kechirasiz...*\n\nArizangiz rad etildi. Admin bilan bog'laning.", parse_mode="Markdown")
        except: pass

broadcast_states = {}

@bot.message_handler(func=lambda m: m.text == "📢 Xabar yuborish" and is_admin(m.from_user.id))
def broadcast_start(message):
    broadcast_states[message.from_user.id] = True
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("🔙 Bekor qilish")
    bot.send_message(message.chat.id, "📢 *Broadcast*\n\nXabarni kiriting:", parse_mode="Markdown", reply_markup=markup)

@bot.message_handler(func=lambda m: m.from_user.id in broadcast_states and broadcast_states.get(m.from_user.id))
def broadcast_send(message):
    user_id = message.from_user.id
    if message.text == "🔙 Bekor qilish":
        broadcast_states.pop(user_id, None)
        bot.send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=main_menu(user_id))
        return
    broadcast_states.pop(user_id, None)
    all_users = db.get_all_user_ids()
    success, failed = 0, 0
    for uid in all_users:
        try:
            bot.send_message(uid, f"📢 *Opus o'quv markazi:*\n\n{message.text}", parse_mode="Markdown")
            success += 1
        except: failed += 1
    bot.send_message(message.chat.id, f"📊 *Natija:*\n✅ {success} ta\n❌ {failed} ta", parse_mode="Markdown", reply_markup=main_menu(user_id))

course_admin_states = {}

@bot.message_handler(func=lambda m: m.text == "📚 Kurslar" and is_admin(m.from_user.id))
def manage_courses(message):
    courses = db.get_courses_with_price()
    text = "📚 *Mavjud kurslar:*\n\n"
    if courses:
        for i, c in enumerate(courses, 1):
            price = f"{c['price']:,} so'm" if c['price'] else "—"
            text += f"{i}. {c['name']} | {price}\n"
    else:
        text += "_Kurslar yo'q_\n"
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("➕ Kurs qo'shish", callback_data="add_course"))
    if courses:
        markup.add(types.InlineKeyboardButton("💰 Narx o'rnatish", callback_data="set_price"))
        markup.add(types.InlineKeyboardButton("🗑 Kurs o'chirish", callback_data="delete_course"))
    bot.send_message(message.chat.id, text, parse_mode="Markdown", reply_markup=markup)

@bot.callback_query_handler(func=lambda c: c.data == "add_course")
def add_course_start(call):
    if not is_admin(call.from_user.id): return
    course_admin_states[call.from_user.id] = "adding"
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, "➕ Yangi kurs nomini kiriting:")

@bot.callback_query_handler(func=lambda c: c.data == "set_price")
def set_price_start(call):
    if not is_admin(call.from_user.id): return
    courses = db.get_all_courses()
    markup = types.InlineKeyboardMarkup()
    for c in courses:
        markup.add(types.InlineKeyboardButton(f"💰 {c['name']}", callback_data=f"price_course_{c['id']}"))
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, "Qaysi kursga narx o'rnatilsin?", reply_markup=markup)

@bot.callback_query_handler(func=lambda c: c.data.startswith("price_course_"))
def set_price_input(call):
    if not is_admin(call.from_user.id): return
    course_id = int(call.data.split("_")[-1])
    course_admin_states[call.from_user.id] = f"pricing_{course_id}"
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, "💰 Narxni kiriting (so'mda, faqat raqam):\n_(Masalan: 500000)_", parse_mode="Markdown")

@bot.callback_query_handler(func=lambda c: c.data == "delete_course")
def delete_course_start(call):
    if not is_admin(call.from_user.id): return
    courses = db.get_all_courses()
    markup = types.InlineKeyboardMarkup()
    for c in courses:
        markup.add(types.InlineKeyboardButton(f"🗑 {c['name']}", callback_data=f"del_course_{c['id']}"))
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, "Qaysi kursni o'chirmoqchisiz?", reply_markup=markup)

@bot.callback_query_handler(func=lambda c: c.data.startswith("del_course_"))
def delete_course_confirm(call):
    if not is_admin(call.from_user.id): return
    course_id = int(call.data.split("_")[-1])
    course = db.get_course_by_id(course_id)
    db.delete_course(course_id)
    bot.answer_callback_query(call.id, "✅ O'chirildi!")
    bot.send_message(call.message.chat.id, f"🗑 '{course['name']}' o'chirildi.")

@bot.message_handler(func=lambda m: m.from_user.id in course_admin_states)
def course_admin_input(message):
    user_id = message.from_user.id
    state = course_admin_states.get(user_id)
    if not state: return
    if state == "adding":
        course_admin_states.pop(user_id, None)
        db.add_course(message.text.strip())
        bot.send_message(message.chat.id, f"✅ '{message.text.strip()}' qo'shildi!", reply_markup=main_menu(user_id))
    elif state.startswith("pricing_"):
        course_id = int(state.split("_")[1])
        course_admin_states.pop(user_id, None)
        try:
            price = int(message.text.strip().replace(" ", "").replace(",", ""))
            db.set_course_price(course_id, price)
            bot.send_message(message.chat.id, f"✅ Narx saqlandi: {price:,} so'm", reply_markup=main_menu(user_id))
        except:
            bot.send_message(message.chat.id, "⚠️ Faqat raqam kiriting!", reply_markup=main_menu(user_id))

schedule_states = {}

@bot.message_handler(func=lambda m: m.text == "📅 Jadval boshqaruv" and is_admin(m.from_user.id))
def manage_schedule(message):
    schedules = db.get_all_schedules()
    text = "📅 *Dars jadvali:*\n\n"
    if schedules:
        for s in schedules:
            text += f"📚 {s['course']} | 👨‍🏫 {s['teacher']}\n🎯 {s['level']} | 📆 {s['days']} | 🕐 {s['time']}\n\n"
    else:
        text += "_Jadval bo'sh_\n"
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("➕ Jadval qo'shish", callback_data="add_schedule"))
    if schedules:
        markup.add(types.InlineKeyboardButton("🗑 Jadval o'chirish", callback_data="del_schedule"))
    bot.send_message(message.chat.id, text, parse_mode="Markdown", reply_markup=markup)

@bot.callback_query_handler(func=lambda c: c.data == "add_schedule")
def add_schedule_start(call):
    if not is_admin(call.from_user.id): return
    schedule_states[call.from_user.id] = {"step": "course", "data": {}}
    bot.answer_callback_query(call.id)
    courses = db.get_all_courses()
    markup = types.InlineKeyboardMarkup()
    for c in courses:
        markup.add(types.InlineKeyboardButton(c['name'], callback_data=f"sch_course_{c['name']}"))
    bot.send_message(call.message.chat.id, "Qaysi kurs uchun jadval?", reply_markup=markup)

@bot.callback_query_handler(func=lambda c: c.data.startswith("sch_course_"))
def sch_course_selected(call):
    if not is_admin(call.from_user.id): return
    course_name = call.data.replace("sch_course_", "")
    schedule_states[call.from_user.id] = {"step": "teacher", "data": {"course": course_name}}
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, f"Kurs: *{course_name}*\n\n👨‍🏫 Ustoz ism-familiyasini kiriting:", parse_mode="Markdown")

@bot.message_handler(func=lambda m: m.from_user.id in schedule_states and schedule_states[m.from_user.id].get("step") in ["teacher","level","days","time"])
def handle_schedule_steps(message):
    user_id = message.from_user.id
    state = schedule_states[user_id]
    step = state["step"]
    text = message.text.strip()
    if step == "teacher":
        state["data"]["teacher"] = text
        state["step"] = "level"
        bot.send_message(message.chat.id, "🎯 Guruh darajasini kiriting:\n_(Masalan: Boshlang'ich, O'rta, Yuqori)_", parse_mode="Markdown")
    elif step == "level":
        state["data"]["level"] = text
        state["step"] = "days"
        bot.send_message(message.chat.id, "📆 Dars kunlarini kiriting:\n_(Masalan: Du-Chor-Ju yoki Se-Pay-Sha)_", parse_mode="Markdown")
    elif step == "days":
        state["data"]["days"] = text
        state["step"] = "time"
        bot.send_message(message.chat.id, "🕐 Dars vaqtini kiriting:\n_(Masalan: 09:00-11:00)_", parse_mode="Markdown")
    elif step == "time":
        state["data"]["time"] = text
        data = state["data"]
        schedule_states.pop(user_id, None)
        db.add_schedule(course=data["course"], teacher=data["teacher"], level=data["level"], days=data["days"], time=data["time"])
        bot.send_message(message.chat.id,
            f"✅ *Jadval qo'shildi!*\n\n📚 {data['course']}\n👨‍🏫 {data['teacher']}\n🎯 {data['level']}\n📆 {data['days']}\n🕐 {data['time']}",
            parse_mode="Markdown", reply_markup=main_menu(user_id))

@bot.callback_query_handler(func=lambda c: c.data == "del_schedule")
def del_schedule_list(call):
    if not is_admin(call.from_user.id): return
    schedules = db.get_all_schedules()
    markup = types.InlineKeyboardMarkup()
    for s in schedules:
        markup.add(types.InlineKeyboardButton(f"🗑 {s['teacher']} | {s['course']} | {s['days']}", callback_data=f"del_sch_{s['id']}"))
    bot.answer_callback_query(call.id)
    bot.send_message(call.message.chat.id, "Qaysi jadvalni o'chirish kerak?", reply_markup=markup)

@bot.callback_query_handler(func=lambda c: c.data.startswith("del_sch_"))
def del_schedule_confirm(call):
    if not is_admin(call.from_user.id): return
    sch_id = int(call.data.split("_")[-1])
    db.delete_schedule(sch_id)
    bot.answer_callback_query(call.id, "✅ O'chirildi!")
    bot.send_message(call.message.chat.id, "🗑 Jadval o'chirildi.")

@bot.message_handler(func=lambda m: m.text == "📤 Excel eksport" and is_admin(m.from_user.id))
def export_excel(message):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        bot.send_message(message.chat.id, "⚠️ Terminalda: pip install openpyxl")
        return
    students = db.get_all_students()
    if not students:
        bot.send_message(message.chat.id, "📭 O'quvchilar yo'q.")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Oqchilar"
    headers = ["#", "Ism-familiya", "Telefon", "Yosh/sinf", "Kurs", "Holat", "Sana"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    status_map = {"pending": "Kutilmoqda", "approved": "Tasdiqlangan", "rejected": "Rad etilgan"}
    for i, s in enumerate(students, 1):
        ws.append([i, s['name'], s['phone'], s['age'], s['course'],
                   status_map.get(s['status'], s['status']),
                   s['created_at'][:10] if s['created_at'] else ""])
    for col, width in zip("ABCDEFG", [4, 25, 16, 12, 18, 14, 12]):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    bot.send_document(message.chat.id, document=buf, visible_file_name="opus_oquvchilar.xlsx",
                      caption=f"📊 Jami {len(students)} ta o'quvchi")

if __name__ == "__main__":
    db.init_db()
    logger.info("Bot ishga tushdi...")
    import threading
from http.server import HTTPServer, BaseHTTPRequestHandler

class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"Bot ishlayapti!")
    def log_message(self, format, *args):
        pass

def run_server():
    HTTPServer(("0.0.0.0", 8080), Handler).serve_forever()

threading.Thread(target=run_server, daemon=True).start()
bot.infinity_polling()
